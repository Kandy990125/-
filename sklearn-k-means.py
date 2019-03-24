#!/usr/bin/python
# -*-coding:utf-8-*-
import sqlite3
import openpyxl
from snownlp import SnowNLP
import requests
import json
def city_lng_lat(city):
    baiduUrl = "http://api.map.baidu.com/geocoder/v2/?ak=S93AVaDxcBpOrPMGhoiLtnAGOxFnrCUQ&callback=renderOption&address=%s&output=json" % (city)
    req = requests.get(baiduUrl)
    content = req.text
    content = content.replace("renderOption&&renderOption(", "")
    content = content[:-1]
    baiduAddr = json.loads(content)
    lng = baiduAddr["result"]["location"]["lng"]
    lat = baiduAddr["result"]["location"]["lat"]
    return lng, lat
def get_snlp(content):
    s = SnowNLP(content)
    i = 0
    sum = 0
    for sentence in s.sentences:
        s = SnowNLP(sentence)
        i = i + 1
        sum = sum + s.sentiments
    avg = sum / i
    return avg

def get_data():
    conn = sqlite3.connect("weibo-spider-scrapy/weibo.sqlite3")
    cursor = conn.cursor()
    exc = "select name from sqlite_master where type='table' order by name"
    list = cursor.execute(exc).fetchall()
    name_list = []
    value_data = []
    for item in list:
        exc = "select user_name,user_sex,user_stay_place,content from %s"%(item)
        item_list = cursor.execute(exc).fetchall()
        for item_data in item_list:
            if item_data[0] in name_list:
                continue
            if len(item_data[1])==0 or len(item_data[2])==0 or item_data[2]=="其他":
                continue
            dir = {}
            if item_data[1]=="男":
                dir["sex"] = 1
            if item_data[1]=="女":
                dir["sex"] = 0
            dir["name"] = item_data[0]
            dir["place"] = item_data[2]
            dir["content"] = item_data[3]
            try:
                dir["snlp"] = get_snlp(item_data[3])
            except:
                dir["snlp"] = -1
            value_data.append(dir)
            name_list.append(dir["name"])
    return value_data

def store(value_list):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet()
    i = 1
    for item in value_list:
        ws.cell(row=i, column=1).value = item["name"]
        ws.cell(row=i, column=2).value = item["sex"]
        try:
            print(i)
            lng, lat = city_lng_lat(item["place"])
        except:
            print("ERROR",i,item["place"])
            lng = 0
            lat = 0
        ws.cell(row=i, column=3).value = lng
        ws.cell(row=i, column=4).value = lat
        ws.cell(row=i, column=5).value = item["content"]
        ws.cell(row=i, column=6).value = item["snlp"]
        ws.cell(row=i, column=7).value = item["place"]
        i = i + 1
    wb.save('data.xlsx')
if __name__=="__main__":
    store(get_data())