#!/usr/bin/python
# -*-coding:utf-8-*-

from openpyxl import load_workbook
from sklearn.cluster import KMeans
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import numpy as np
import requests
import json
def get_city(lng,lat):
    url = "http://api.map.baidu.com/geocoder/v2/?ak=S93AVaDxcBpOrPMGhoiLtnAGOxFnrCUQ&callback=renderReverse&location=%s&output=json&pois=1"%(str(lat)+","+str(lng))
    req = requests.get(url)
    content = req.text
    content = content.replace("renderReverse&&renderReverse(", "")
    content = content[:-1]
    dir = json.loads(content)
    province = dir["result"]["addressComponent"]["province"]
    city = dir["result"]["addressComponent"]["city"]
    return province,city
def get_data():
    wb = load_workbook(filename=r'data1.xlsx')
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    x = []
    y = []
    for rx in range(2, ws.max_row + 1):
        pid = rx
        w2 = ws.cell(row=rx, column=1).value
        w3 = ws.cell(row=rx, column=2).value
        x.append(w2)
        y.append(w3)
    return x,y
def basic_plot(x,y):
    plt.rc('font', family='STXihei', size=10)
    plt.scatter(x, y, 50, color='blue', marker='+', linewidth=2, alpha=0.8)
    plt.xlabel('x')
    plt.ylabel('y')
    plt.grid(color='#95a5a6', linestyle='--', linewidth=1, axis='both', alpha=0.4)
    plt.show()
def kmeans_plot():
    wb = openpyxl.Workbook()
    ws = wb.create_sheet()
    loan_data = pd.DataFrame(pd.read_excel('data.xlsx', header=0))
    # 设置要进行聚类的字段
    loan = np.array(loan_data[['x', 'y']])
    # 设置类别为3
    clf = KMeans(n_clusters=8)
    # 将数据代入到聚类模型中
    clf = clf.fit(loan)
    centers = np.array(clf.cluster_centers_)
    i = 1
    for center in centers:
        lng = center[0]
        lat = center[1]
        province,city = get_city(lng,lat)
        ws.cell(row=i, column=1).value = lng
        ws.cell(row=i, column=2).value = lat
        ws.cell(row=i, column=3).value = province
        ws.cell(row=i, column=4).value = city
        i = i + 1
    loan_data['label'] = clf.labels_
    # loan_data0 = loan_data.loc[loan_data["label"] == 0]
    #     loan_data1 = loan_data.loc[loan_data["label"] == 1]
    #     loan_data2 = loan_data.loc[loan_data["label"] == 2]
    #     loan_data3 = loan_data.loc[loan_data["label"] == 3]
    #     loan_data4 = loan_data.loc[loan_data["label"] == 4]
    #     loan_data5 = loan_data.loc[loan_data["label"] == 5]
    #     loan_data6 = loan_data.loc[loan_data["label"] == 6]
    #     loan_data7 = loan_data.loc[loan_data["label"] == 7]
    # plt.rc('font', family='STXihei', size=10)
    #     plt.scatter(loan_data0['x'], loan_data0['y'], 50, color='#99CC01', marker='+', linewidth=2,alpha=0.8)
    #     plt.scatter(loan_data1['x'], loan_data1['y'], 50, color='#FE0000', marker='+', linewidth=2,alpha=0.8)
    #     plt.scatter(loan_data2['x'], loan_data2['y'], 50, color='#0000FE', marker='+', linewidth=2,alpha=0.8)
    #     plt.scatter(loan_data3['x'], loan_data3['y'], 50, color='#D15FEE', marker='+', linewidth=2,alpha=0.8)
    #     plt.scatter(loan_data4['x'], loan_data4['y'], 50, color='#EE7600', marker='+', linewidth=2,alpha=0.8)
    #     plt.scatter(loan_data5['x'], loan_data5['y'], 50, color='#FF82AB', marker='+', linewidth=2,alpha=0.8)
    #     plt.scatter(loan_data6['x'], loan_data6['y'], 50, color='#B0E2FF', marker='+', linewidth=2,alpha=0.8)
    #     plt.scatter(loan_data7['x'], loan_data7['y'], 50, color='yellow', marker='+', linewidth=2, alpha=0.8)
    #     plt.scatter([i[0] for i in centers], [i[1] for i in centers], 50, color='black', marker='*', linewidth=2, alpha=0.8)
    #     plt.xlabel('x')
    #     plt.ylabel('y')
    #     plt.grid(color='#95a5a6', linestyle='--', linewidth=1, axis='both', alpha=0.4)
    #     plt.show()
    wb.save('data3.xlsx')
    return loan_data

def main1():
    loan_data = np.array(kmeans_plot())
    wb = openpyxl.Workbook()
    ws = wb.create_sheet()
    i = 1
    for item in loan_data:
        ws.cell(row=i, column=1).value = item[0]
        ws.cell(row=i, column=2).value = item[1]
        ws.cell(row=i, column=3).value = item[2]
        ws.cell(row=i, column=4).value = item[3]
        ws.cell(row=i, column=5).value = item[4]
        ws.cell(row=i, column=6).value = item[5]
        ws.cell(row=i, column=7).value = item[6]
        ws.cell(row=i, column=8).value = item[7]
        i = i + 1

    wb.save('data2.xlsx')

main1()