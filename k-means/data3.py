#!/usr/bin/python
# -*-coding:utf-8-*-

from openpyxl import load_workbook
from sklearn.cluster import KMeans
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import numpy as np
# 绿色长三角地区label=0
def store(value_list):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet()
    rx = 1
    for dir in value_list:
        ws.cell(row=rx, column=1).value =dir["name"]
        ws.cell(row=rx, column=2).value = dir["sex"]
        ws.cell(row=rx, column=3).value = dir["lng"]
        ws.cell(row=rx, column=4).value =dir["lat"]
        ws.cell(row=rx, column=5).value=dir["content"]
        ws.cell(row=rx, column=6).value=dir["snlp"]
        rx = rx + 1
    wb.save('label7.xlsx')

def get_data():
    wb = load_workbook(filename=r'data2.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')
    array = []
    for rx in range(1, ws.max_row + 1):
        label = int(ws.cell(row=rx, column=7).value)
        if label==7:
            dir = {}
            dir["name"] = ws.cell(row=rx, column=1).value
            dir["sex"] = ws.cell(row=rx,column=2).value
            dir["lng"] = ws.cell(row=rx,column=3).value
            dir["lat"] = ws.cell(row=rx,column=4).value
            dir["content"] = ws.cell(row=rx,column=5).value
            dir["snlp"] = ws.cell(row=rx,column=6).value
            array.append(dir)
    return array
store(get_data())